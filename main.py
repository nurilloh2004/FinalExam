from operator import le
import sys
import time
import traceback
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from models import Flat, Row
from windows.MarketWindow import MarketWindow
from windows.FlatWindow import FlatWindow
from openpyxl import Workbook


class Window(QMainWindow):

    def __init__(self) -> None:
        super().__init__()

        self.initUI()
        self.initActions()
        self.initMenu()
        self.initTable()
        self.fillTable()
        self.msg = QMessageBox()

    def initUI(self):
        btn_add = QPushButton("Add", self)
        btn_add.move(900, 720)
        btn_add.clicked.connect(self.onAdd)

        btn_add = QPushButton("Update", self)
        btn_add.move(1020, 720)
        btn_add.clicked.connect(self.onUpdate)

        btn_add = QPushButton("Delete", self)
        btn_add.move(1140, 720)
        btn_add.clicked.connect(self.onDel)

        btn_add = QPushButton("Report", self)
        btn_add.move(1260, 720)
        btn_add.clicked.connect(self.onRep)

        ql = QLabel("Product_name: ", self)
        ql.move(1450, 95)
        ql = QLabel("Company_name: ", self)
        ql.move(1450, 125)
        ql = QLabel("Expiration_date: ", self)
        ql.move(1450, 155)
        ql = QLabel("Price: ", self)
        ql.move(1450, 185)
        ql = QLabel("Discount_price: ", self)
        ql.move(1450, 215)
        ql = QLabel("FlatId: ", self)
        ql.move(1450, 245)

        self.qle_product_name = QLineEdit(self)
        self.qle_product_name.move(1550, 95)
        self.qle_company_name = QLineEdit(self)
        self.qle_company_name.move(1550, 125)
        self.qle_expiration_date = QLineEdit(self)
        self.qle_expiration_date.move(1550, 155)
        self.qle_price = QLineEdit(self)
        self.qle_price.move(1550, 185)
        self.qle_discount_price = QLineEdit(self)
        self.qle_discount_price.move(1550, 215)
        self.cbb_flat = QComboBox(self)
        self.cbb_flat.move(1550, 245)
        for item in Flat.objects():
            self.cbb_flat.addItem(item.name, item.id)

    def onRep(self):
        wb = Workbook()
        try:
            # grab the active worksheet
            ws = wb.active
            ws[f'A{1}'] = "Product_name"
            ws[f'B{1}'] = "Company_name"
            ws[f'C{1}'] = "Expiration_date'ulgan yili"
            ws[f'D{1}'] = "Price"
            ws[f'E{1}'] = "Discount_price"
            ws[f'F{1}'] = "Markets"
            ws[f'G{1}'] = "Flats"

            for sel_row in range(self.table.rowCount()):
                product_name = self.table.item(sel_row, 1).text()
                company_name = self.table.item(sel_row, 2).text()
                expiration_date = int(self.table.item(sel_row, 3).text())
                price = int(self.table.item(sel_row, 4).text())
                discount_price = int(self.table.item(sel_row, 5).text())
                mar_name = self.table.item(sel_row, 6).text()
                flat_name = self.table.item(sel_row, 8).text()

                ws[f'A{sel_row + 2}'] = product_name
                ws[f'B{sel_row + 2}'] = company_name
                ws[f'C{sel_row + 2}'] = expiration_date
                ws[f'D{sel_row + 2}'] = price
                ws[f'E{sel_row + 2}'] = discount_price
                ws[f'F{sel_row + 2}'] = mar_name
                ws[f'G{sel_row + 2}'] = flat_name
            # Save the file
            wb.save("sample.xlsx")
            wb.close()

        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()
            wb.close()
            traceback.print_exc()

    def onAdd(self):
        try:
            product_name = self.qle_product_name.text()
            company_name = self.qle_company_name.text()
            expiration_date = int(self.qle_expiration_date.text())
            price = int(self.qle_price.text())
            discount_price = int(self.qle_discount_price.text())
            flatId = self.cbb_flat.currentData()

            row = Row(product_name, company_name, expiration_date, price, discount_price, flatId)
            row.save()

            flat = row.flat
            mar = flat.market
            row_count = self.table.rowCount()
            self.table.setRowCount(row_count + 1)
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(row.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(row.product_name))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(row.company_name))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(row.expiration_date)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(row.price)))
            self.table.setItem(row_count, 5,
                               QTableWidgetItem(str(row.discount_price)))

            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(mar.name)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(str(flat.id)))
            self.table.setItem(row_count, 8,
                               QTableWidgetItem(flat.name))
            self.msg.setIcon(QMessageBox.Information)
            self.msg.setWindowTitle("Bajarildi")
            self.msg.setText("Talaba saqlandi....")
            self.msg.show()
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def onUpdate(self):
        try:
            id = int(self.table.item(self.sel_row, 0).text())
            product_name = self.qle_product_name.text()
            company_name = self.qle_company_name.text()
            expiration_date = int(self.qle_expiration_date.text())
            price = int(self.qle_price.text())
            discount_price = int(self.qle_discount_price.text())
            flatId = self.cbb_flat.currentData()

            row = Row(product_name, company_name, expiration_date, price, discount_price, flatId, id)
            row.save()

            flat = row.flat
            mar = flat.market
            row_count = self.table.currentRow()
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(row.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(row.product_name))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(row.company_name))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(row.expiration_date)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(row.price)))
            self.table.setItem(row_count, 5,
                               QTableWidgetItem(str(row.discount_price)))

            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(mar.name)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(str(flat.id)))
            self.table.setItem(row_count, 8,
                               QTableWidgetItem(flat.name))
            self.msg.setIcon(QMessageBox.Information)
            self.msg.setWindowTitle("Bajarildi")
            self.msg.setText("Talaba saqlandi....")
            self.msg.show()
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def onDel(self):
        try:
            id = int(self.table.item(self.sel_row, 0).text())
            product_name = self.qle_product_name.text()
            company_name = self.qle_company_name.text()
            expiration_date = int(self.qle_expiration_date.text())
            price = int(self.qle_price.text())
            discount_price = int(self.qle_discount_price.text())
            flatId = self.cbb_flat.currentData()

            row = Row(product_name, company_name, expiration_date, price, discount_price, flatId, id)
            row.delete()

            id = int(self.table.item(self.sel_row, 0).text())
            self.qle_product_name.setText('')
            self.qle_company_name.setText('')
            self.qle_expiration_date.setText('')
            self.qle_price.setText('')
            self.qle_discount_price.setText('')

            row_count = self.table.currentRow()
            self.table.removeRow(row_count)

            self.msg.setIcon(QMessageBox.Information)
            self.msg.setWindowTitle("Bajarildi")
            self.msg.setText("Talaba o'chirildi....")
            self.msg.show()
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def initTable(self):
        self.table = QTableWidget(self)  # Создаём таблицу
        self.table.move(400, 100)
        self.table.setMinimumSize(1000, 600)
        self.table.setColumnCount(9)     # Устанавливаем три колонки
        self.table.setSelectionBehavior(QTableWidget.SelectRows)

        # Устанавливаем заголовки таблицы
        self.table.setHorizontalHeaderLabels(
            ['Id', "Product_name", "Company_name", "Expiration_date", "Price", "Discount_price", "Markets", "Flat Id", "Flat"])

        self.table.hideColumn(0)
        self.table.hideColumn(7)
        self.table.clicked.connect(self.onClicked)

    def initActions(self):
        self.newAction = QAction("&New...", self)
        self.newAction.triggered.connect(self.onnewAction)
        self.openAction = QAction("&Open...", self)
        self.saveAction = QAction("&Save", self)
        self.exitAction = QAction("&Exit", self)

        self.marketAction = QAction("&Markets", self)
        self.marketAction.triggered.connect(self.onMarketWindow)
        self.flatAction = QAction("&Flats", self)
        self.flatAction.triggered.connect(self.onFlatWindow)

    def fillTable(self):

        for row in Row.objects():
            flat = row.flat
            mar = flat.market
            row_count = self.table.rowCount()
            self.table.setRowCount(row_count + 1)
            self.table.setItem(row_count, 0,
                               QTableWidgetItem(str(row.id)))
            self.table.setItem(row_count, 1,
                               QTableWidgetItem(row.product_name))
            self.table.setItem(row_count, 2,
                               QTableWidgetItem(row.company_name))
            self.table.setItem(row_count, 3,
                               QTableWidgetItem(str(row.expiration_date)))
            self.table.setItem(row_count, 4,
                               QTableWidgetItem(str(row.price)))
            self.table.setItem(row_count, 5,
                               QTableWidgetItem(str(row.discount_price)))

            self.table.setItem(row_count, 6,
                               QTableWidgetItem(str(mar.name)))
            self.table.setItem(row_count, 7,
                               QTableWidgetItem(str(flat.id)))
            self.table.setItem(row_count, 8,
                               QTableWidgetItem(flat.name))
            # делаем ресайз колонок по содержимому
        self.table.resizeColumnsToContents()

    def onClicked(self):
        try:
            self.sel_row = self.table.currentRow()
            product_name = self.table.item(self.sel_row, 1).text()
            self.qle_product_name.setText(product_name)
            company_name = self.table.item(self.sel_row, 2).text()
            self.qle_company_name.setText(company_name)
            expiration_date = int(self.table.item(self.sel_row, 3).text())
            self.qle_expiration_date.setText(str(expiration_date))
            price = int(self.table.item(self.sel_row, 4).text())
            self.qle_price.setText(str(price))
            discount_price = int(self.table.item(self.sel_row, 5).text())
            self.qle_discount_price.setText(str(discount_price))
            flatId = int(self.table.item(self.sel_row, 7).text())
            for i in range(self.cbb_flat.count()):
                if self.cbb_flat.itemData(i) == flatId:
                    self.cbb_flat.setCurrentIndex(i)
                    break
        except Exception as exp:
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setWindowTitle("Xatolik")
            self.msg.setText(str(exp))
            self.msg.show()

            traceback.print_exc()

    def onnewAction(self):
        pass

    def initMenu(self):
        menuBar = self.menuBar()
        self.setMenuBar(menuBar)

        fileMenu = menuBar.addMenu("&File")
        fileMenu.addAction(self.newAction)
        fileMenu.addAction(self.openAction)
        fileMenu.addAction(self.saveAction)
        fileMenu.addAction(self.exitAction)

        servicesMenu = menuBar.addMenu("&Services")
        servicesMenu.addAction(self.marketAction)
        servicesMenu.addAction(self.flatAction)

        helpMenu = menuBar.addMenu("&Help")

    def onMarketWindow(self):
        self.marw = MarketWindow()
        self.marw.show()

    def onFlatWindow(self):
        self.flatw = FlatWindow()
        self.flatw.show()


app = QApplication(sys.argv)

w = Window()
w.showMaximized()

app.exec()
